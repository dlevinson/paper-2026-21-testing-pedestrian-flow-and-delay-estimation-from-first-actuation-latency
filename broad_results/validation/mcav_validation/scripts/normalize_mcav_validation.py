#!/usr/bin/env python3
"""Normalize the Manual Crossing-Actuation Validation (MCAV) workbook.

The source workbook is field-observation data, not SCATS HST data.  This
script treats it as independent validation evidence for first-actuation
pedestrian-volume inference.
"""

from __future__ import annotations

import csv
import json
import math
import random
import re
import statistics
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[4]
SOURCE_XLSX = (
    ROOT
    / "Pedestrian Actuators Validation Project"
    / "Manual Crossing-Actuation Validation"
    / "11_Intersections_Data.xlsx"
)
OUTPUT_DIR = ROOT / "broad_results" / "validation" / "mcav_validation"

CYCLE_OUT = OUTPUT_DIR / "mcav_cycle_level_validation.csv"
CASE_OUT = OUTPUT_DIR / "mcav_case_model_validation.csv"
SUMMARY_OUT = OUTPUT_DIR / "mcav_model_comparison_summary.csv"
STABILITY_OUT = OUTPUT_DIR / "mcav_model_stability_summary.csv"
LOO_OUT = OUTPUT_DIR / "mcav_model_leave_one_out.csv"
PLOT_JSON_OUT = OUTPUT_DIR / "mcav_model_plot_data.json"
COMPLIANCE_CASE_OUT = OUTPUT_DIR / "mcav_compliance_scenario_by_case.csv"
COMPLIANCE_SUMMARY_OUT = OUTPUT_DIR / "mcav_compliance_scenario_summary.csv"


CELL_REF_RE = re.compile(r"^='(?P<sheet>.+)'!(?P<col>[A-Z]+)(?P<row>\d+)$")
SUM_RANGE_RE = re.compile(r"^=SUM\([A-Z]+(?P<start>\d+):[A-Z]+(?P<end>\d+)\)$")


def number(value) -> float:
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        if isinstance(value, float) and math.isnan(value):
            return 0.0
        return float(value)
    return 0.0


def text(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def parse_cell_ref(formula: str) -> tuple[str, int]:
    match = CELL_REF_RE.match(text(formula))
    if not match:
        raise ValueError(f"Cannot parse source cell reference: {formula!r}")
    return match.group("sheet"), int(match.group("row"))


def parse_sum_range(formula: str) -> tuple[int, int]:
    match = SUM_RANGE_RE.match(text(formula))
    if not match:
        raise ValueError(f"Cannot parse source SUM range: {formula!r}")
    return int(match.group("start")), int(match.group("end"))


def is_walk(phase: str) -> bool:
    phase = phase.strip().upper()
    # Workbook rows labelled ``W (Invalid)`` occur before the first usable
    # cycle anchor and are not valid Walk exposure.
    return phase == "W"


def is_nonwalk(phase: str) -> bool:
    phase = phase.strip().upper()
    return phase.startswith("FDW") or phase.startswith("SDW")


def read_cases(wb_formula, wb_values) -> list[dict]:
    formula_sheet = wb_formula["Comparison "]
    value_sheet = wb_values["Comparison "]
    cases = []
    row = 2
    while text(formula_sheet.cell(row, 1).value):
        case = text(formula_sheet.cell(row, 1).value)
        source_sheet, source_total_row = parse_cell_ref(formula_sheet.cell(row, 2).value)
        formula_ws = wb_formula[source_sheet]
        data_start, data_end = parse_sum_range(formula_ws.cell(source_total_row, 3).value)
        cases.append(
            {
                "case": case,
                "comparison_row": row,
                "source_sheet": source_sheet,
                "source_total_row": source_total_row,
                "data_start_row": data_start,
                "data_end_row": data_end,
                "comparison_legal_pushbutton": number(value_sheet.cell(row, 2).value),
                "comparison_legal_no_pushbutton": number(value_sheet.cell(row, 3).value),
                "comparison_illegal_pushbutton": number(value_sheet.cell(row, 4).value),
                "comparison_illegal_no_pushbutton": number(value_sheet.cell(row, 5).value),
                "comparison_pushbutton_no_walk": number(value_sheet.cell(row, 6).value),
                "comparison_actual_total": number(value_sheet.cell(row, 7).value),
                "comparison_workbook_predicted_total": number(value_sheet.cell(row, 8).value),
                "comparison_mean_seconds_fdw_to_first_push": number(value_sheet.cell(row, 10).value),
                "workbook_manual_actuator_count": number(
                    wb_values[source_sheet].cell(source_total_row, 4).value
                ),
            }
        )
        row += 1
    return cases


def read_case_records(case: dict, wb_values) -> list[dict]:
    ws = wb_values[case["source_sheet"]]
    records = []
    expected_time = 1
    anomaly_count = 0
    for row in range(case["data_start_row"], case["data_end_row"] + 1):
        time_value = ws.cell(row, 1).value
        if isinstance(time_value, (int, float)) and int(time_value) != expected_time:
            anomaly_count += 1
        records.append(
            {
                "row": row,
                "case_second": expected_time,
                "raw_time_seconds": time_value if isinstance(time_value, (int, float)) else "",
                "phase": text(ws.cell(row, 2).value),
                "ped_arrivals": number(ws.cell(row, 3).value),
                "actuator_count_raw": number(ws.cell(row, 4).value),
                "legal_pushbutton": number(ws.cell(row, 5).value),
                "legal_no_pushbutton": number(ws.cell(row, 6).value),
                "illegal_pushbutton": number(ws.cell(row, 7).value),
                "illegal_no_pushbutton": number(ws.cell(row, 8).value),
                "pushbutton_no_walk": number(ws.cell(row, 9).value),
            }
        )
        expected_time += 1
    case["observed_seconds"] = len(records)
    case["raw_time_anomaly_count"] = anomaly_count
    return records


def sum_records(records: list[dict], start_idx: int, end_idx: int, key: str) -> float:
    return sum(record[key] for record in records[start_idx:end_idx])


def cycle_rows_for_case(case: dict, records: list[dict]) -> list[dict]:
    nonwalk_starts = []
    previous_nonwalk = False
    for idx, record in enumerate(records):
        current_nonwalk = is_nonwalk(record["phase"])
        if current_nonwalk and not previous_nonwalk:
            nonwalk_starts.append(idx)
        previous_nonwalk = current_nonwalk

    rows = []
    for cycle_number, start_idx in enumerate(nonwalk_starts, start=1):
        next_start_idx = (
            nonwalk_starts[cycle_number]
            if cycle_number < len(nonwalk_starts)
            else len(records)
        )

        walk_start_idx = None
        for idx in range(start_idx + 1, next_start_idx):
            if is_walk(records[idx]["phase"]):
                walk_start_idx = idx
                break

        nonwalk_end_idx = walk_start_idx if walk_start_idx is not None else next_start_idx
        first_push_idx = None
        for idx in range(start_idx, nonwalk_end_idx):
            if records[idx]["actuator_count_raw"] > 0:
                first_push_idx = idx
                break

        start_record = records[start_idx]
        next_start_record = records[next_start_idx] if next_start_idx < len(records) else None
        walk_start_record = records[walk_start_idx] if walk_start_idx is not None else None
        first_push_record = records[first_push_idx] if first_push_idx is not None else None

        nonwalk_exposure = nonwalk_end_idx - start_idx
        full_cycle_exposure = next_start_idx - start_idx
        first_push_wait = (
            first_push_idx - start_idx if first_push_idx is not None else ""
        )
        effective_wait = (
            max(float(first_push_wait), 0.5)
            if first_push_idx is not None
            else ""
        )

        if effective_wait != "":
            predicted_nonwalk = nonwalk_exposure / effective_wait
            predicted_full = full_cycle_exposure / effective_wait
        else:
            predicted_nonwalk = 0.0
            predicted_full = 0.0

        rows.append(
            {
                "case": case["case"],
                "source_sheet": case["source_sheet"],
                "comparison_row": case["comparison_row"],
                "source_total_row": case["source_total_row"],
                "data_start_row": case["data_start_row"],
                "data_end_row": case["data_end_row"],
                "cycle_number": cycle_number,
                "nonwalk_start_row": start_record["row"],
                "nonwalk_start_second": start_record["case_second"],
                "walk_start_row": walk_start_record["row"] if walk_start_record else "",
                "walk_start_second": walk_start_record["case_second"] if walk_start_record else "",
                "next_nonwalk_start_row": (
                    next_start_record["row"] if next_start_record else ""
                ),
                "next_nonwalk_start_second": (
                    next_start_record["case_second"] if next_start_record else ""
                ),
                "first_push_row": first_push_record["row"] if first_push_record else "",
                "first_push_second": (
                    first_push_record["case_second"] if first_push_record else ""
                ),
                "first_push_observed": 1 if first_push_record else 0,
                "first_push_wait_seconds": first_push_wait,
                "effective_wait_seconds": effective_wait,
                "nonwalk_exposure_seconds": nonwalk_exposure,
                "full_cycle_exposure_seconds": full_cycle_exposure,
                "observed_ped_arrivals_cycle": sum_records(
                    records, start_idx, next_start_idx, "ped_arrivals"
                ),
                "observed_ped_arrivals_nonwalk": sum_records(
                    records, start_idx, nonwalk_end_idx, "ped_arrivals"
                ),
                "actuator_count_raw_cycle": sum_records(
                    records, start_idx, next_start_idx, "actuator_count_raw"
                ),
                "legal_pushbutton_cycle": sum_records(
                    records, start_idx, next_start_idx, "legal_pushbutton"
                ),
                "legal_no_pushbutton_cycle": sum_records(
                    records, start_idx, next_start_idx, "legal_no_pushbutton"
                ),
                "illegal_pushbutton_cycle": sum_records(
                    records, start_idx, next_start_idx, "illegal_pushbutton"
                ),
                "illegal_no_pushbutton_cycle": sum_records(
                    records, start_idx, next_start_idx, "illegal_no_pushbutton"
                ),
                "pushbutton_no_walk_cycle": sum_records(
                    records, start_idx, next_start_idx, "pushbutton_no_walk"
                ),
                "cycle_plugin_predicted_nonwalk_only": predicted_nonwalk,
                "cycle_plugin_predicted_full_cycle": predicted_full,
            }
        )
    return rows


def aggregate_case(case: dict, records: list[dict], cycles: list[dict]) -> dict:
    actual = sum(record["ped_arrivals"] for record in records)
    raw_actuator_half_sum = sum(record["actuator_count_raw"] for record in records) / 2.0
    manual_actuator = case["workbook_manual_actuator_count"]
    first_push_cycles = sum(row["first_push_observed"] for row in cycles)
    known_cycles = len(cycles)
    no_push_cycles = known_cycles - first_push_cycles
    actual_first_push = sum(
        row["observed_ped_arrivals_cycle"]
        for row in cycles
        if row["first_push_observed"]
    )
    actual_no_push = actual - actual_first_push
    wait_sum = sum(
        float(row["first_push_wait_seconds"])
        for row in cycles
        if row["first_push_observed"]
    )
    censored_nonwalk = sum(
        row["nonwalk_exposure_seconds"]
        for row in cycles
        if not row["first_push_observed"]
    )
    total_risk = wait_sum + censored_nonwalk
    total_nonwalk = sum(row["nonwalk_exposure_seconds"] for row in cycles)
    total_full = sum(row["full_cycle_exposure_seconds"] for row in cycles)
    observed_seconds = len(records)

    poisson_lambda = first_push_cycles / total_risk if total_risk > 0 else 0.0
    observed_only_lambda = first_push_cycles / wait_sum if wait_sum > 0 else 0.0

    pooled_observed_first_push = observed_only_lambda * observed_seconds
    pooled_censored_all_time = poisson_lambda * observed_seconds
    pooled_censored_nonwalk = poisson_lambda * total_nonwalk
    predicted_nonwalk = sum(row["cycle_plugin_predicted_nonwalk_only"] for row in cycles)
    predicted_full = sum(row["cycle_plugin_predicted_full_cycle"] for row in cycles)

    raw_totals = {
        "legal_pushbutton": sum(record["legal_pushbutton"] for record in records),
        "legal_no_pushbutton": sum(record["legal_no_pushbutton"] for record in records),
        "illegal_pushbutton": sum(record["illegal_pushbutton"] for record in records),
        "illegal_no_pushbutton": sum(record["illegal_no_pushbutton"] for record in records),
        "pushbutton_no_walk": sum(record["pushbutton_no_walk"] for record in records),
    }

    return {
        "case": case["case"],
        "source_sheet": case["source_sheet"],
        "data_start_row": case["data_start_row"],
        "data_end_row": case["data_end_row"],
        "observed_seconds": observed_seconds,
        "raw_time_anomaly_count": case["raw_time_anomaly_count"],
        "actual_total_pedestrians": actual,
        "comparison_actual_total": case["comparison_actual_total"],
        "manual_actuator_count": manual_actuator,
        "raw_actuator_half_sum": raw_actuator_half_sum,
        "pedestrians_per_manual_actuation": (
            actual / manual_actuator if manual_actuator else ""
        ),
        "legal_pushbutton": raw_totals["legal_pushbutton"],
        "legal_no_pushbutton": raw_totals["legal_no_pushbutton"],
        "illegal_pushbutton": raw_totals["illegal_pushbutton"],
        "illegal_no_pushbutton": raw_totals["illegal_no_pushbutton"],
        "pushbutton_no_walk": raw_totals["pushbutton_no_walk"],
        "known_nonwalk_intervals": known_cycles,
        "first_push_intervals": first_push_cycles,
        "censored_no_push_intervals": no_push_cycles,
        "actual_pedestrians_in_first_push_cycles": actual_first_push,
        "actual_pedestrians_in_no_push_cycles": actual_no_push,
        "sum_first_push_wait_seconds": wait_sum,
        "sum_censored_nonwalk_seconds": censored_nonwalk,
        "total_risk_exposure_seconds": total_risk,
        "total_nonwalk_seconds": total_nonwalk,
        "total_full_cycle_seconds": total_full,
        "poisson_observed_first_push_implied_count_all_time": pooled_observed_first_push,
        "poisson_censored_implied_count_all_time": pooled_censored_all_time,
        "poisson_censored_implied_count_nonwalk_only": pooled_censored_nonwalk,
        "workbook_first_push_predicted_total": case[
            "comparison_workbook_predicted_total"
        ],
        "cycle_plugin_predicted_nonwalk_only": predicted_nonwalk,
        "cycle_plugin_predicted_full_cycle": predicted_full,
    }


def compliance_scenario_case(
    case: dict, records: list[dict], cycles: list[dict], aggregate: dict
) -> dict:
    """Evaluate universal signal and actuator compliance in one MCAV case.

    Universal signal compliance assigns every arrival during a closed interval
    the time remaining to the following observed Walk. Universal actuator
    participation moves the first actuation in each closed interval to the
    first observed pedestrian arrival, while preserving the estimator and
    censoring rules. This separates missing actuation from batching and the
    remaining timing/model mismatch.
    """

    counterfactual_first_arrival_intervals = 0
    counterfactual_risk_seconds = 0.0
    full_compliance_delay_seconds = 0.0
    delay_covered_restrictive_arrivals = 0.0
    delay_uncovered_restrictive_arrivals = 0.0
    occupied_arrival_seconds = sum(record["ped_arrivals"] > 0 for record in records)
    pedestrians_on_multiarrival_seconds = sum(
        record["ped_arrivals"] for record in records if record["ped_arrivals"] > 1
    )
    observed_model_delay_unexpanded_seconds = 0.0
    first_actuation_model_delay_unexpanded_seconds = 0.0
    cycle_plugin_delay_unexpanded_seconds = 0.0
    full_actuation_model_delay_unexpanded_seconds = 0.0

    observed_lambda = (
        aggregate["first_push_intervals"] / aggregate["total_risk_exposure_seconds"]
        if aggregate["total_risk_exposure_seconds"]
        else 0.0
    )
    first_actuation_lambda = (
        aggregate["first_push_intervals"] / aggregate["sum_first_push_wait_seconds"]
        if aggregate["sum_first_push_wait_seconds"]
        else 0.0
    )

    first_arrivals: list[tuple[dict, int | None, int, int]] = []
    for cycle in cycles:
        start_idx = int(cycle["nonwalk_start_second"]) - 1
        if cycle["walk_start_second"] != "":
            nonwalk_end_idx = int(cycle["walk_start_second"]) - 1
        elif cycle["next_nonwalk_start_second"] != "":
            nonwalk_end_idx = int(cycle["next_nonwalk_start_second"]) - 1
        else:
            nonwalk_end_idx = len(records)

        first_arrival_idx = next(
            (
                idx
                for idx in range(start_idx, nonwalk_end_idx)
                if records[idx]["ped_arrivals"] > 0
            ),
            None,
        )
        first_arrivals.append((cycle, first_arrival_idx, start_idx, nonwalk_end_idx))
        if first_arrival_idx is None:
            counterfactual_risk_seconds += nonwalk_end_idx - start_idx
        else:
            counterfactual_first_arrival_intervals += 1
            counterfactual_risk_seconds += first_arrival_idx - start_idx

    full_actuation_lambda = (
        counterfactual_first_arrival_intervals / counterfactual_risk_seconds
        if counterfactual_risk_seconds
        else 0.0
    )
    full_actuation_implied_count = full_actuation_lambda * len(records)

    for cycle, first_arrival_idx, start_idx, nonwalk_end_idx in first_arrivals:
        if cycle["walk_start_second"] == "":
            delay_uncovered_restrictive_arrivals += sum(
                records[idx]["ped_arrivals"]
                for idx in range(start_idx, nonwalk_end_idx)
            )
            continue

        walk_idx = int(cycle["walk_start_second"]) - 1
        for idx in range(start_idx, walk_idx):
            pedestrians = records[idx]["ped_arrivals"]
            delay_covered_restrictive_arrivals += pedestrians
            full_compliance_delay_seconds += pedestrians * (walk_idx - idx)

        if cycle["first_push_second"] != "":
            first_push_idx = int(cycle["first_push_second"]) - 1
            residual = walk_idx - first_push_idx
            observed_model_delay_unexpanded_seconds += (
                residual + observed_lambda * residual**2 / 2
            )
            first_actuation_model_delay_unexpanded_seconds += (
                residual + first_actuation_lambda * residual**2 / 2
            )
            cycle_rate = 1.0 / float(cycle["effective_wait_seconds"])
            cycle_plugin_delay_unexpanded_seconds += (
                residual + cycle_rate * residual**2 / 2
            )
        if first_arrival_idx is not None:
            residual = walk_idx - first_arrival_idx
            full_actuation_model_delay_unexpanded_seconds += (
                residual + full_actuation_lambda * residual**2 / 2
            )

    actual = aggregate["actual_total_pedestrians"]
    observed_implied_count = aggregate["poisson_censored_implied_count_all_time"]
    first_actuation_implied_count = aggregate[
        "poisson_observed_first_push_implied_count_all_time"
    ]
    cycle_plugin_implied_count = aggregate["cycle_plugin_predicted_full_cycle"]
    count_alpha_observed = actual / observed_implied_count
    count_alpha_first_actuation = actual / first_actuation_implied_count
    count_alpha_full_actuation = actual / full_actuation_implied_count
    signal_noncompliant = aggregate["illegal_pushbutton"] + aggregate["illegal_no_pushbutton"]
    observed_crossers = (
        aggregate["legal_pushbutton"]
        + aggregate["legal_no_pushbutton"]
        + signal_noncompliant
    )
    compliance_share = (
        1.0 - signal_noncompliant / observed_crossers
        if observed_crossers
        else 1.0
    )
    # Behaviour classes are recorded as cycle totals rather than linked to
    # individual arrival timestamps.  This scenario therefore applies the
    # case's observed compliant share to its full-compliance delay, assuming
    # non-compliers would otherwise have the same expected wait and incur zero
    # signal delay when they cross.  It is not a direct realised-delay measure.
    observed_compliance_delay_seconds_scenario = (
        full_compliance_delay_seconds * compliance_share
    )
    noncompliance_delay_reduction_seconds_scenario = (
        full_compliance_delay_seconds
        - observed_compliance_delay_seconds_scenario
    )
    return {
        "case": case["case"],
        "actual_total_pedestrians": actual,
        "observed_crossers": observed_crossers,
        "signal_noncompliant_crossers": signal_noncompliant,
        "signal_noncompliance_rate": (
            signal_noncompliant / observed_crossers if observed_crossers else ""
        ),
        "observed_actuator_implied_count": observed_implied_count,
        "count_alpha_observed_actuation": count_alpha_observed,
        "first_actuation_latency_implied_count": first_actuation_implied_count,
        "count_alpha_first_actuation_latency": count_alpha_first_actuation,
        "cycle_plugin_full_cycle_implied_count": cycle_plugin_implied_count,
        "count_alpha_cycle_plugin_full_cycle": actual / cycle_plugin_implied_count,
        "actual_pedestrians_in_first_push_cycles": aggregate[
            "actual_pedestrians_in_first_push_cycles"
        ],
        "actual_pedestrians_in_no_push_cycles": aggregate[
            "actual_pedestrians_in_no_push_cycles"
        ],
        "full_actuation_first_arrival_intervals": counterfactual_first_arrival_intervals,
        "full_actuation_risk_seconds": counterfactual_risk_seconds,
        "full_actuation_implied_count": full_actuation_implied_count,
        "count_alpha_full_actuation": count_alpha_full_actuation,
        "missing_actuation_multiplier": (
            full_actuation_implied_count / observed_implied_count
        ),
        "occupied_arrival_seconds": occupied_arrival_seconds,
        "pedestrians_on_multiarrival_seconds": pedestrians_on_multiarrival_seconds,
        "batch_factor_people_per_occupied_second": actual / occupied_arrival_seconds,
        "remaining_timing_model_factor": (
            occupied_arrival_seconds / full_actuation_implied_count
        ),
        "full_compliance_delay_seconds": full_compliance_delay_seconds,
        "observed_compliance_delay_seconds_scenario": (
            observed_compliance_delay_seconds_scenario
        ),
        "noncompliance_delay_reduction_seconds_scenario": (
            noncompliance_delay_reduction_seconds_scenario
        ),
        "delay_covered_restrictive_arrivals": delay_covered_restrictive_arrivals,
        "delay_uncovered_restrictive_arrivals": delay_uncovered_restrictive_arrivals,
        "walk_or_zero_delay_arrivals": (
            actual
            - delay_covered_restrictive_arrivals
            - delay_uncovered_restrictive_arrivals
        ),
        "average_full_compliance_delay_all_pedestrians_seconds_lower_bound": (
            full_compliance_delay_seconds / actual if actual else ""
        ),
        "average_full_compliance_delay_restrictive_seconds": (
            full_compliance_delay_seconds / delay_covered_restrictive_arrivals
            if delay_covered_restrictive_arrivals
            else ""
        ),
        "observed_actuation_model_delay_unexpanded_seconds": (
            observed_model_delay_unexpanded_seconds
        ),
        "delay_alpha_observed_actuation": (
            full_compliance_delay_seconds / observed_model_delay_unexpanded_seconds
            if observed_model_delay_unexpanded_seconds
            else ""
        ),
        "delay_alpha_observed_compliance_scenario": (
            observed_compliance_delay_seconds_scenario
            / observed_model_delay_unexpanded_seconds
            if observed_model_delay_unexpanded_seconds
            else ""
        ),
        "first_actuation_model_delay_unexpanded_seconds": (
            first_actuation_model_delay_unexpanded_seconds
        ),
        "delay_alpha_first_actuation_latency": (
            full_compliance_delay_seconds
            / first_actuation_model_delay_unexpanded_seconds
            if first_actuation_model_delay_unexpanded_seconds
            else ""
        ),
        "cycle_plugin_delay_unexpanded_seconds": (
            cycle_plugin_delay_unexpanded_seconds
        ),
        "delay_alpha_cycle_plugin_full_cycle": (
            full_compliance_delay_seconds / cycle_plugin_delay_unexpanded_seconds
            if cycle_plugin_delay_unexpanded_seconds
            else ""
        ),
        "full_actuation_model_delay_unexpanded_seconds": (
            full_actuation_model_delay_unexpanded_seconds
        ),
        "delay_alpha_full_actuation": (
            full_compliance_delay_seconds / full_actuation_model_delay_unexpanded_seconds
            if full_actuation_model_delay_unexpanded_seconds
            else ""
        ),
        "observed_count_alpha_scaled_model_delay_seconds": (
            count_alpha_observed * observed_model_delay_unexpanded_seconds
        ),
    }


def compliance_scenario_summary(rows: list[dict]) -> list[dict]:
    actual = sum(row["actual_total_pedestrians"] for row in rows)
    actual_in_first_push_cycles = sum(
        row["actual_pedestrians_in_first_push_cycles"] for row in rows
    )
    actual_in_no_push_cycles = sum(
        row["actual_pedestrians_in_no_push_cycles"] for row in rows
    )
    observed_implied = sum(row["observed_actuator_implied_count"] for row in rows)
    first_actuation_implied = sum(
        row["first_actuation_latency_implied_count"] for row in rows
    )
    cycle_plugin_implied = sum(
        row["cycle_plugin_full_cycle_implied_count"] for row in rows
    )
    full_actuation_implied = sum(row["full_actuation_implied_count"] for row in rows)
    occupied_seconds = sum(row["occupied_arrival_seconds"] for row in rows)
    full_delay = sum(row["full_compliance_delay_seconds"] for row in rows)
    observed_compliance_delay = sum(
        row["observed_compliance_delay_seconds_scenario"] for row in rows
    )
    noncompliance_delay_reduction = sum(
        row["noncompliance_delay_reduction_seconds_scenario"] for row in rows
    )
    observed_model_delay = sum(
        row["observed_actuation_model_delay_unexpanded_seconds"] for row in rows
    )
    first_actuation_model_delay = sum(
        row["first_actuation_model_delay_unexpanded_seconds"] for row in rows
    )
    cycle_plugin_model_delay = sum(
        row["cycle_plugin_delay_unexpanded_seconds"] for row in rows
    )
    full_actuation_model_delay = sum(
        row["full_actuation_model_delay_unexpanded_seconds"] for row in rows
    )
    observed_count_alphas = [row["count_alpha_observed_actuation"] for row in rows]
    first_actuation_count_alphas = [
        row["count_alpha_first_actuation_latency"] for row in rows
    ]
    cycle_plugin_count_alphas = [
        row["count_alpha_cycle_plugin_full_cycle"] for row in rows
    ]
    full_actuation_count_alphas = [row["count_alpha_full_actuation"] for row in rows]
    observed_delay_alphas = [row["delay_alpha_observed_actuation"] for row in rows]
    observed_compliance_delay_alphas = [
        row["delay_alpha_observed_compliance_scenario"] for row in rows
    ]
    first_actuation_delay_alphas = [
        row["delay_alpha_first_actuation_latency"] for row in rows
    ]
    cycle_plugin_delay_alphas = [
        row["delay_alpha_cycle_plugin_full_cycle"] for row in rows
    ]
    full_actuation_delay_alphas = [row["delay_alpha_full_actuation"] for row in rows]
    signal_noncompliant = sum(row["signal_noncompliant_crossers"] for row in rows)
    observed_crossers = sum(row["observed_crossers"] for row in rows)
    delay_covered = sum(row["delay_covered_restrictive_arrivals"] for row in rows)
    delay_uncovered = sum(row["delay_uncovered_restrictive_arrivals"] for row in rows)
    multiarrival_pedestrians = sum(
        row["pedestrians_on_multiarrival_seconds"] for row in rows
    )
    mean_observed_count_alpha = statistics.fmean(observed_count_alphas)
    mean_observed_delay_alpha = statistics.fmean(observed_delay_alphas)
    mean_first_actuation_count_alpha = statistics.fmean(
        first_actuation_count_alphas
    )
    mean_first_actuation_delay_alpha = statistics.fmean(
        first_actuation_delay_alphas
    )
    mean_cycle_plugin_count_alpha = statistics.fmean(cycle_plugin_count_alphas)
    mean_cycle_plugin_delay_alpha = statistics.fmean(cycle_plugin_delay_alphas)
    metrics = {
        "case_count": len(rows),
        "actual_total_pedestrians": actual,
        "observed_crossers": observed_crossers,
        "signal_noncompliant_crossers": signal_noncompliant,
        "signal_noncompliance_rate": signal_noncompliant / observed_crossers,
        "observed_actuator_implied_count": observed_implied,
        "count_alpha_observed_ratio_of_sums": actual / observed_implied,
        "count_alpha_observed_mean_case": mean_observed_count_alpha,
        "first_actuation_latency_implied_count": first_actuation_implied,
        "count_alpha_first_actuation_ratio_of_sums": actual
        / first_actuation_implied,
        "count_alpha_first_actuation_mean_case": mean_first_actuation_count_alpha,
        "count_alpha_first_actuation_min": min(first_actuation_count_alphas),
        "count_alpha_first_actuation_max": max(first_actuation_count_alphas),
        "cycle_plugin_full_cycle_implied_count": cycle_plugin_implied,
        "count_alpha_cycle_plugin_ratio_of_sums": actual / cycle_plugin_implied,
        "count_alpha_cycle_plugin_mean_case": mean_cycle_plugin_count_alpha,
        "count_alpha_cycle_plugin_min": min(cycle_plugin_count_alphas),
        "count_alpha_cycle_plugin_max": max(cycle_plugin_count_alphas),
        "actual_pedestrians_in_first_push_cycles": actual_in_first_push_cycles,
        "actual_pedestrians_in_no_push_cycles": actual_in_no_push_cycles,
        "no_push_cycle_coverage_factor": actual / actual_in_first_push_cycles,
        "within_first_push_cycle_factor": actual_in_first_push_cycles
        / cycle_plugin_implied,
        "cycle_plugin_decomposition_product_check": (
            (actual / actual_in_first_push_cycles)
            * (actual_in_first_push_cycles / cycle_plugin_implied)
        ),
        "full_actuation_implied_count": full_actuation_implied,
        "count_alpha_full_actuation_ratio_of_sums": actual / full_actuation_implied,
        "count_alpha_full_actuation_mean_case": statistics.fmean(
            full_actuation_count_alphas
        ),
        "count_alpha_full_actuation_min": min(full_actuation_count_alphas),
        "count_alpha_full_actuation_max": max(full_actuation_count_alphas),
        "missing_actuation_multiplier_ratio_of_sums": (
            full_actuation_implied / observed_implied
        ),
        "occupied_arrival_seconds": occupied_seconds,
        "pedestrians_on_multiarrival_seconds": multiarrival_pedestrians,
        "pedestrian_share_on_multiarrival_seconds": multiarrival_pedestrians / actual,
        "batch_factor_people_per_occupied_second": actual / occupied_seconds,
        "remaining_timing_model_factor": occupied_seconds / full_actuation_implied,
        "decomposition_product_check": (
            (full_actuation_implied / observed_implied)
            * (actual / occupied_seconds)
            * (occupied_seconds / full_actuation_implied)
        ),
        "full_compliance_delay_seconds": full_delay,
        "full_compliance_delay_hours": full_delay / 3600,
        "observed_compliance_delay_seconds_scenario": observed_compliance_delay,
        "observed_compliance_delay_hours_scenario": observed_compliance_delay / 3600,
        "noncompliance_delay_reduction_seconds_scenario": (
            noncompliance_delay_reduction
        ),
        "noncompliance_delay_reduction_hours_scenario": (
            noncompliance_delay_reduction / 3600
        ),
        "observed_compliance_to_full_compliance_delay_ratio": (
            observed_compliance_delay / full_delay
        ),
        "delay_covered_restrictive_arrivals": delay_covered,
        "delay_uncovered_restrictive_arrivals": delay_uncovered,
        "average_full_compliance_delay_all_pedestrians_seconds_lower_bound": (
            full_delay / actual
        ),
        "average_full_compliance_delay_restrictive_seconds": full_delay / delay_covered,
        "observed_actuation_model_delay_unexpanded_seconds": observed_model_delay,
        "delay_alpha_observed_actuation_ratio_of_sums": full_delay / observed_model_delay,
        "delay_alpha_observed_actuation_mean_case": mean_observed_delay_alpha,
        "delay_alpha_observed_actuation_min": min(observed_delay_alphas),
        "delay_alpha_observed_actuation_max": max(observed_delay_alphas),
        "delay_alpha_observed_compliance_ratio_of_sums": (
            observed_compliance_delay / observed_model_delay
        ),
        "delay_alpha_observed_compliance_mean_case": statistics.fmean(
            observed_compliance_delay_alphas
        ),
        "delay_alpha_observed_compliance_median_case": statistics.median(
            observed_compliance_delay_alphas
        ),
        "delay_alpha_observed_compliance_min": min(
            observed_compliance_delay_alphas
        ),
        "delay_alpha_observed_compliance_max": max(
            observed_compliance_delay_alphas
        ),
        "first_actuation_model_delay_unexpanded_seconds": (
            first_actuation_model_delay
        ),
        "delay_alpha_first_actuation_ratio_of_sums": full_delay
        / first_actuation_model_delay,
        "delay_alpha_first_actuation_mean_case": mean_first_actuation_delay_alpha,
        "delay_alpha_first_actuation_min": min(first_actuation_delay_alphas),
        "delay_alpha_first_actuation_max": max(first_actuation_delay_alphas),
        "cycle_plugin_delay_unexpanded_seconds": cycle_plugin_model_delay,
        "delay_alpha_cycle_plugin_ratio_of_sums": full_delay
        / cycle_plugin_model_delay,
        "delay_alpha_cycle_plugin_mean_case": mean_cycle_plugin_delay_alpha,
        "delay_alpha_cycle_plugin_min": min(cycle_plugin_delay_alphas),
        "delay_alpha_cycle_plugin_max": max(cycle_plugin_delay_alphas),
        "full_actuation_model_delay_unexpanded_seconds": full_actuation_model_delay,
        "delay_alpha_full_actuation_ratio_of_sums": (
            full_delay / full_actuation_model_delay
        ),
        "delay_alpha_full_actuation_mean_case": statistics.fmean(
            full_actuation_delay_alphas
        ),
        "delay_using_mean_count_alpha_hours": (
            observed_model_delay * mean_observed_count_alpha / 3600
        ),
        "delay_using_mean_delay_alpha_hours": (
            observed_model_delay * mean_observed_delay_alpha / 3600
        ),
        "first_actuation_delay_using_mean_alpha_hours": (
            first_actuation_model_delay * mean_first_actuation_delay_alpha / 3600
        ),
        "cycle_plugin_delay_using_mean_alpha_hours": (
            cycle_plugin_model_delay * mean_cycle_plugin_delay_alpha / 3600
        ),
    }
    return [{"metric": key, "value": value} for key, value in metrics.items()]


def abs_relative_error(actual: float, predicted: float) -> float:
    if actual == 0:
        return 0.0 if predicted == 0 else math.inf
    return abs(predicted - actual) / actual


def write_csv(path: Path, rows: list[dict], fieldnames: list[str] | None = None) -> None:
    if fieldnames is None:
        fieldnames = list(rows[0].keys()) if rows else []
    with path.open("w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def model_summary(case_rows: list[dict]) -> list[dict]:
    models = [
        (
            "manual_actuator_count",
            "Diagnostic recorded actuator events, not a person-volume estimator",
        ),
        (
            "workbook_first_push_predicted_total",
            "Workbook first-wait formula, 3600 divided by mean first-push wait",
        ),
        (
            "poisson_observed_first_push_implied_count_all_time",
            "Pooled first-push Poisson, uncensored first-push cycles only",
        ),
        (
            "poisson_censored_implied_count_all_time",
            "Pooled first-push Poisson with no-push cycles as right-censored exposure",
        ),
        (
            "poisson_censored_implied_count_nonwalk_only",
            "Pooled first-push Poisson with censored exposure, non-walk time only",
        ),
        (
            "cycle_plugin_predicted_nonwalk_only",
            "Cycle plug-in first-actuation model, non-walk exposure only",
        ),
        (
            "cycle_plugin_predicted_full_cycle",
            "Cycle plug-in first-actuation model, full observed cycle exposure",
        ),
    ]

    total_actual = sum(row["actual_total_pedestrians"] for row in case_rows)
    rows = []
    for field, label in models:
        total_predicted = sum(float(row[field]) for row in case_rows)
        case_errors = [
            abs_relative_error(row["actual_total_pedestrians"], float(row[field]))
            for row in case_rows
        ]
        rows.append(
            {
                "model_key": field,
                "model": label,
                "actual_total_pedestrians": total_actual,
                "implied_total_pedestrians": total_predicted,
                "implied_to_actual_ratio": (
                    total_predicted / total_actual if total_actual else ""
                ),
                "calibration_multiplier_actual_over_implied": (
                    total_actual / total_predicted if total_predicted else ""
                ),
                "mean_absolute_relative_error_cases": (
                    sum(case_errors) / len(case_errors) if case_errors else ""
                ),
                "median_absolute_relative_error_cases": (
                    sorted(case_errors)[len(case_errors) // 2] if case_errors else ""
                ),
            }
        )
    return rows


def stability(case_rows: list[dict]) -> tuple[list[dict], list[dict], dict]:
    model_fields = [
        (
            "poisson_observed_first_push_implied_count_all_time",
            "First-actuation latency, observed first-actuation phases",
        ),
        (
            "poisson_censored_implied_count_all_time",
            "Pooled first-push Poisson with no-push cycles as right-censored exposure",
        ),
        (
            "cycle_plugin_predicted_nonwalk_only",
            "Cycle plug-in first-actuation model, non-walk exposure only",
        ),
        (
            "cycle_plugin_predicted_full_cycle",
            "Cycle plug-in first-actuation model, full observed cycle exposure",
        ),
    ]
    total_actual = sum(row["actual_total_pedestrians"] for row in case_rows)
    summary_rows = []
    loo_rows = []
    plot_data = {"cases": [], "summary": []}

    for row in case_rows:
        plot_data["cases"].append(
            {
                "case": row["case"],
                "actual": row["actual_total_pedestrians"],
                "poisson_censored_all_time": row[
                    "poisson_censored_implied_count_all_time"
                ],
                "nonwalk": row["cycle_plugin_predicted_nonwalk_only"],
                "full_cycle": row["cycle_plugin_predicted_full_cycle"],
                "pushbutton_share": (
                    (row["legal_pushbutton"] + row["illegal_pushbutton"])
                    / row["actual_total_pedestrians"]
                    if row["actual_total_pedestrians"]
                    else None
                ),
            }
        )

    rng = random.Random(20260710)
    for field, label in model_fields:
        predicted = sum(row[field] for row in case_rows)
        aggregate_ratio = predicted / total_actual
        loo_ratios = []
        for left_out in case_rows:
            actual_minus = total_actual - left_out["actual_total_pedestrians"]
            predicted_minus = predicted - left_out[field]
            ratio = predicted_minus / actual_minus if actual_minus else math.nan
            loo_ratios.append(ratio)
            loo_rows.append(
                {
                    "model_key": field,
                    "model": label,
                    "left_out_case": left_out["case"],
                    "implied_to_actual_ratio_leave_one_out": ratio,
                }
            )

        bootstrap_ratios = []
        n = len(case_rows)
        for _ in range(10000):
            sample = [case_rows[rng.randrange(n)] for _ in range(n)]
            sample_actual = sum(row["actual_total_pedestrians"] for row in sample)
            sample_predicted = sum(row[field] for row in sample)
            bootstrap_ratios.append(sample_predicted / sample_actual)
        bootstrap_ratios.sort()
        lo = bootstrap_ratios[int(0.025 * len(bootstrap_ratios))]
        hi = bootstrap_ratios[int(0.975 * len(bootstrap_ratios))]
        summary_rows.append(
            {
                "model_key": field,
                "model": label,
                "aggregate_implied_to_actual_ratio": aggregate_ratio,
                "aggregate_calibration_multiplier": 1 / aggregate_ratio,
                "leave_one_out_min_ratio": min(loo_ratios),
                "leave_one_out_max_ratio": max(loo_ratios),
                "bootstrap_ratio_p025": lo,
                "bootstrap_ratio_p975": hi,
                "bootstrap_multiplier_p025": 1 / hi,
                "bootstrap_multiplier_p975": 1 / lo,
            }
        )
        plot_data["summary"].append(summary_rows[-1])
    return summary_rows, loo_rows, plot_data


def main() -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    wb_formula = load_workbook(SOURCE_XLSX, data_only=False, read_only=False)
    wb_values = load_workbook(SOURCE_XLSX, data_only=True, read_only=False)

    cases = read_cases(wb_formula, wb_values)
    all_cycles = []
    case_rows = []
    compliance_rows = []
    for case in cases:
        records = read_case_records(case, wb_values)
        cycles = cycle_rows_for_case(case, records)
        all_cycles.extend(cycles)
        aggregate = aggregate_case(case, records, cycles)
        case_rows.append(aggregate)
        compliance_rows.append(
            compliance_scenario_case(case, records, cycles, aggregate)
        )

    write_csv(CYCLE_OUT, all_cycles)
    write_csv(CASE_OUT, case_rows)
    write_csv(SUMMARY_OUT, model_summary(case_rows))
    stability_rows, loo_rows, plot_data = stability(case_rows)
    write_csv(STABILITY_OUT, stability_rows)
    write_csv(LOO_OUT, loo_rows)
    write_csv(COMPLIANCE_CASE_OUT, compliance_rows)
    write_csv(COMPLIANCE_SUMMARY_OUT, compliance_scenario_summary(compliance_rows))
    PLOT_JSON_OUT.write_text(json.dumps(plot_data, indent=2), encoding="utf-8")

    actual = sum(row["actual_total_pedestrians"] for row in case_rows)
    actuations = sum(row["manual_actuator_count"] for row in case_rows)
    cycles = len(all_cycles)
    first_push = sum(row["first_push_observed"] for row in all_cycles)
    print(f"Cases: {len(case_rows)}")
    print(f"Cycles: {cycles} ({first_push} with first push)")
    print(f"Observed pedestrians: {actual:.0f}")
    print(f"Manual actuator events: {actuations:.0f}")
    print(f"Wrote: {CYCLE_OUT}")
    print(f"Wrote: {CASE_OUT}")
    print(f"Wrote: {SUMMARY_OUT}")
    print(f"Wrote: {STABILITY_OUT}")
    print(f"Wrote: {LOO_OUT}")
    print(f"Wrote: {COMPLIANCE_CASE_OUT}")
    print(f"Wrote: {COMPLIANCE_SUMMARY_OUT}")
    print(f"Wrote: {PLOT_JSON_OUT}")


if __name__ == "__main__":
    main()
